import datetime
import os

from openpyxl import load_workbook, Workbook

from config import logger, engine_kwargs, robot_name, global_env_data, main_folder
from utils.check_if_excel_good import get_last_excel
from utils.create_infotable_excel import create_infotable_excel

from sqlalchemy import create_engine, Column, Integer, String, DateTime, MetaData, Table, Date, Boolean, select, BigInteger, distinct
from sqlalchemy.orm import declarative_base, sessionmaker


Base = declarative_base()


class IsmetTable(Base):

    __tablename__ = "parse_all"
    id = Column(Integer, primary_key=True)
    # add_time = Column(DateTime, default=None)
    edit_time = Column(DateTime, default=None)
    status = Column(String(128), default=None)
    error_message = Column(String(128), default=None)

    ID_INVOICE = Column(String(128))
    NUMBER_INVOICE = Column(String(128))
    URL_INVOICE = Column(String(128))
    C_NAME_SOURCE_INVOICE = Column(String(128))
    C_NAME_SHOP = Column(String(128))
    DATE_INVOICE = Column(DateTime)
    BAR_CODE_WARES = Column(BigInteger)
    NAME_WARES = Column(String(128))
    QUANTITY = Column(Integer)
    APPROVE_FLAG = Column(Integer, default=None)

    @property
    def dict(self):
        m = self.__dict__
        return m


def check_invoice_in_db():

    """
        Creating connection to the ismet table
        Getting all rows from the ismet table with type(NUMBER_INVOICE) = string
        Creating new excel with only rows with status 'Отклонён'
    """

    # * Creating connection to the ismet table

    Session_ismet = sessionmaker()

    engine_kwargs_ismet = {
        'username': global_env_data['postgre_db_username'],
        'password': global_env_data['postgre_db_password'],
        'host': global_env_data['postgre_ip'],
        'port': global_env_data['postgre_port'],
        'base': 'ismet'
    }

    engine = create_engine(
        'postgresql+psycopg2://{username}:{password}@{host}:{port}/{base}'.format(**engine_kwargs_ismet),
        connect_args={'options': '-csearch_path=public'}
    )

    Base.metadata.create_all(bind=engine)
    Session_ismet.configure(bind=engine)
    session_ismet = Session_ismet()

    # * Getting all rows with status 'Отклонён'

    from_date = datetime.date(2023, 10, 23)
    condition = IsmetTable.edit_time >= from_date

    select_query = (
        session_ismet.query(IsmetTable)
        .filter(condition)
        .filter(IsmetTable.NUMBER_INVOICE.like('!%'))
        .all()
    )

    # * Fetching all number invoices from the db

    invoices = dict()
    for ind, row in enumerate(select_query):
        if row.ID_INVOICE not in list(invoices.keys()):

            invoices.update({row.ID_INVOICE: [row.NUMBER_INVOICE, row.C_NAME_SHOP, row.C_NAME_SOURCE_INVOICE, row.edit_time]})
            # print(a.ID_INVOICE, a.NUMBER_INVOICE, a.DATE_INVOICE, sep=' | ')

    # * Creating new Excel from rows with status 'Отклонён'

    book = Workbook()
    sheet = book.active

    sheet['A1'].value = '№'
    sheet['B1'].value = 'Дата'
    sheet['C1'].value = 'Филиал'
    sheet['D1'].value = 'Поставщик'
    sheet['E1'].value = 'Код отклонения'
    sheet['F1'].value = 'Причина отклонения'

    last_row = sheet.max_row + 1

    for key, val in invoices.items():
        print(key, val)
        sheet[f'A{last_row}'].value = str(key)
        sheet[f'B{last_row}'].value = val[3].strftime('%d.%m.%Y')
        sheet[f'C{last_row}'].value = val[1]
        sheet[f'D{last_row}'].value = val[2]
        sheet[f'E{last_row}'].value = val[0]
        if val[0] == '!DECLINE FORCED':
            sheet[f'F{last_row}'].value = 'Выявлены отклонения после ручной проверки'
        if val[0] == '!DECLINE FOUND BUT INCORRECT QUANTITY':
            sheet[f'F{last_row}'].value = 'Соответствие найдено, но количество позиции или позиций неверно'
        if val[0] == '!DECLINE INVOICE NOT FOUND':
            sheet[f'F{last_row}'].value = 'Нет соответствия в данных  за указанный период'
        if val[0] == '!DECLINE SOURCE NOT FOUND':
            sheet[f'F{last_row}'].value = 'Нет поставщика в данных Магнума за указанный период'
        last_row += 1

    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 100
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 50

    excel_name = f"{datetime.date.today().strftime('%Y-%m-%d')}_ismet_рассылка.xlsx"

    book.save(excel_name)

    return [excel_name, invoices]

# if __name__ == '__main__':
#
#     check_invoice_in_db()


