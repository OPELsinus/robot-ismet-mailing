import os
from time import sleep

import numpy
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config import main_folder, working_path


def divide_excel_by_suppliers(excel_file: str):

    # * Getting count of rows for each supplier

    df = pd.read_excel(excel_file)

    print(df['Поставщик'].unique())

    suppliers = dict()

    for supplier in df['Поставщик'].unique():

        print(supplier, df[df['Поставщик'] == supplier]['Поставщик'])

        df[df['Поставщик'] == supplier].copy().to_excel(os.path.join(working_path, supplier.replace('"', '') + '.xlsx'), index=False)

        sleep(.2)

        book = load_workbook(os.path.join(working_path, supplier.replace('"', '') + '.xlsx'))
        sheet = book.active

        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = len(df[df['Поставщик'] == supplier]['Поставщик'].iloc[0]) * 1.15
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 50

        column_index = sheet['E'][0].column

        sheet.delete_cols(column_index)

        book.save(os.path.join(working_path, supplier.replace('"', '') + '.xlsx'))

        suppliers.update({supplier: os.path.join(working_path, supplier.replace('"', '') + '.xlsx')})

        print('------')

    return suppliers
