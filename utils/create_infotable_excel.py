import os

import numpy
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config import main_folder


def create_infotable_excel(excel_file: str):

    # * Getting count of rows for each supplier

    df = pd.read_excel(excel_file)

    print(df['Поставщик'].unique())

    suppliers = dict()

    for supplier in df['Поставщик'].unique():
        print(supplier, len(df[df['Поставщик'] == supplier]))
        suppliers.update({supplier: len(df[df['Поставщик'] == supplier])})

    # * Creating info table on the new sheet

    book = load_workbook(excel_file)
    sheet = book.create_sheet('Сводная таблица')

    for row, vals in enumerate(suppliers.items()):

        sheet[f'A{row + 2}'].value = vals[0]
        sheet[f'B{row + 2}'].value = vals[1]

    sheet[f'A1'].value = 'Поставщики'
    sheet[f'B1'].value = 'Отклонён'
    sheet[f'A1'].font = Font(bold=True)
    sheet[f'B1'].font = Font(bold=True)

    # * Designing the sheet

    sheet[f'A{len(suppliers) + 2}'].value = 'Общий итог'
    sheet[f'B{len(suppliers) + 2}'].value = sum([i for i in suppliers.values()])
    sheet[f'A{len(suppliers) + 2}'].font = Font(bold=True)
    sheet[f'B{len(suppliers) + 2}'].font = Font(bold=True)

    fill_color = PatternFill(start_color="00CCCCFF", end_color="00CCCCFF", fill_type="solid")

    try:
        sheet.column_dimensions['A'].width = max([len(i) for i in suppliers.keys()]) * 1.15
    except:
        sheet.column_dimensions['A'].width = 15

    sheet.column_dimensions['B'].width = 15
    sheet['B1'].alignment = Alignment(horizontal='center')

    sheet['A1'].fill = fill_color
    sheet['B1'].fill = fill_color
    sheet[f'A{len(suppliers) + 2}'].fill = fill_color
    sheet[f'B{len(suppliers) + 2}'].fill = fill_color

    book.save(excel_file)


