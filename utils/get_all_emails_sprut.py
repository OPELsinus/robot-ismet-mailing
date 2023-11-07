import pyperclip
from openpyxl import load_workbook
from pywinauto import keyboard

from config import mapping_file
from core import Sprut


def get_all_emails(suppliers):

    mapping_excel = load_workbook(mapping_file)

    mapping_sheet = mapping_excel.active

    sprut = Sprut('MAGNUM')
    sprut.run()

    root = sprut.parent

    emails = dict()

    for row in range(2, mapping_sheet.max_row + 1):

        for supplier in suppliers:

            if mapping_sheet[f'A{row}'].value != supplier:
                continue

            print(f'Started {supplier}')

            sprut.open("Физические лица, юридические лица и производители")

            sprut.find_element({"title": "", "class_name": "TvmsDBToolGrid", "control_type": "Pane",
                                "visible_only": True, "enabled_only": True, "found_index": 1}).click()
            sprut.find_element({"title": "", "class_name": "TvmsDBToolGrid", "control_type": "Pane",
                                "visible_only": True, "enabled_only": True, "found_index": 1}).type_keys('^F')
            sprut.parent_switch({"title": "Поиск", "class_name": "Tvms_search_fm_builder", "control_type": "Window",
                                 "visible_only": True, "enabled_only": True, "found_index": 0}).set_focus()

            # * Selecting 'Юридическое лицо'

            sprut.find_element({"title_re": ".", "class_name": "TvmsComboBox", "control_type": "Pane",
                                "visible_only": True, "enabled_only": True, "found_index": 0}).click()
            sprut.find_element({"title_re": ".", "class_name": "TvmsComboBox", "control_type": "Pane",
                                "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys(sprut.keys.UP * 15)
            sprut.find_element({"title_re": ".", "class_name": "TvmsComboBox", "control_type": "Pane",
                                "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys(sprut.keys.DOWN * 3, sprut.keys.ENTER)

            # * Searching for the company

            sprut.find_element({"title": "", "class_name": "TcxCustomInnerTextEdit", "control_type": "Edit",
                                "visible_only": True, "enabled_only": True, "found_index": 0}).click()
            sprut.find_element({"title": "", "class_name": "TcxCustomInnerTextEdit", "control_type": "Edit",
                                "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys('^N')
            sprut.find_element({"title": "", "class_name": "TcxCustomInnerTextEdit", "control_type": "Edit",
                                "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys(mapping_sheet[f'B{row}'].value, sprut.keys.ENTER, protect_first=True)

            # * Clicking

            sprut.find_element({"title": "Выбрать", "class_name": "TvmsBitBtn", "control_type": "Button",
                                "visible_only": True, "enabled_only": True, "found_index": 0}).click()

            sprut.parent_back(1)

            # * Setting cursor at the most left column

            sprut.find_element({"title": "", "class_name": "TvmsDBToolGrid", "control_type": "Pane",
                                "visible_only": True, "enabled_only": True, "found_index": 1}).click()
            sprut.find_element({"title": "", "class_name": "TvmsDBToolGrid", "control_type": "Pane",
                                "visible_only": True, "enabled_only": True, "found_index": 1}).type_keys(sprut.keys.LEFT * 30)

            # * Searching for emails

            for col in range(20):

                keyboard.send_keys('^%{INSERT}')

                column_name = pyperclip.paste()

                if column_name == 'Адрес электронной почты':

                    keyboard.send_keys('^{INSERT}')

                    emails.update({mapping_sheet[f'A{row}'].value: pyperclip.paste()})

                    break

                sprut.find_element({"title": "", "class_name": "TvmsDBToolGrid", "control_type": "Pane",
                                    "visible_only": True, "enabled_only": True, "found_index": 1}).type_keys(sprut.keys.RIGHT)

            sprut.find_element({"title": "Выход", "class_name": "", "control_type": "MenuItem",
                                "visible_only": True, "enabled_only": True, "found_index": 0}).click()

            sprut.parent_switch(root)

            print(f'Finished {supplier}')

    sprut.quit()

    return emails


