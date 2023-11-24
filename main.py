import datetime
from time import sleep

from openpyxl import load_workbook

from config import logger, engine_kwargs, robot_name, smtp_host, smtp_author
from tools.smtp import smtp_send
from utils.check_if_excel_good import get_last_excel
from utils.check_invoice_in_db import check_invoice_in_db
from utils.create_infotable_excel import create_infotable_excel

from sqlalchemy import create_engine, Column, Integer, String, DateTime, MetaData, Table, Date, Boolean, select
from sqlalchemy.orm import declarative_base, sessionmaker

from utils.divide import divide_excel_by_suppliers
from utils.get_all_emails_sprut import get_all_emails

Base = declarative_base()


class Table(Base):

    __tablename__ = robot_name.replace('-', '_')

    date_created = Column(DateTime, default=None)
    invoice_date = Column(DateTime, default=None)
    id_invoice = Column(String(512), primary_key=True)
    reason_invoice = Column(String(512), default=None)
    store_name = Column(String(512), default=None)
    supplier_name = Column(String(512), default=None)

    status = Column(String(16), default=None)

    @property
    def dict(self):
        m = self.__dict__.copy()
        return m


if __name__ == '__main__':

    Session = sessionmaker()

    engine = create_engine(
        'postgresql+psycopg2://{username}:{password}@{host}:{port}/{base}'.format(**engine_kwargs),
        connect_args={'options': '-csearch_path=robot'}
    )
    Base.metadata.create_all(bind=engine)
    Session.configure(bind=engine)
    session = Session()

    excel_file, invoices = check_invoice_in_db()
    print(invoices)
    for key, val in invoices.items():
        # print(key, val)
        # print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), val[3])
        session.add(Table(
            date_created=datetime.datetime.now(),
            invoice_date=val[3],
            id_invoice=key,
            reason_invoice=val[0],
            store_name=val[1],
            supplier_name=val[2],
            status='new'
        ))
    session.commit()

    suppliers_excels: dict = divide_excel_by_suppliers(excel_file)

    create_infotable_excel(excel_file)

    emails = get_all_emails(suppliers_excels.keys())
    # emails = {'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "ALMA TRADE DISTRIBUTION"': 'fortisline.elnar@mail.ru, uchet.fortis@mail.ru, akty.almatrade@gmail.com', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "FORTIS SKO"': 'rogacheva.1981@mail.ru', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "DITRADE KRG"': 'ditradekaraganda@mail.ru, ditradekrg@mail.ru', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "ТОРГОВАЯ КОМПАНИЯ "МЕГАПОЛИС-КАЗАХСТАН"': 'zemlyanukhin.daniil@gkm-kz.com, nikolai_kireev_89@mail.ru, megapolis_kam@mail.ru, megapolis.redbull@gmail.com, chshepkin@gkm-kz.com, golovin.sanya.71@gmail.com, shaihiev.i@outlook.com, TKmegapolis.zakaz@gkm-kz.com', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "CITY TRADE AST"': 'ast_city_trade@mail.ru', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "KAZNORD"': 'buh3009@mail.ru'}
    # emails = {'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "ALMA TRADE DISTRIBUTION"': 'fortisline.elnar@mail.ru, uchet.fortis@mail.ru, akty.almatrade@gmail.com', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "FORTIS SKO"': 'rogacheva.1981@mail.ru', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "DITRADE KRG"': 'ditradekaraganda@mail.ru, ditradekrg@mail.ru', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "ТОРГОВАЯ КОМПАНИЯ "МЕГАПОЛИС-КАЗАХСТАН"': 'zemlyanukhin.daniil@gkm-kz.com, nikolai_kireev_89@mail.ru, megapolis_kam@mail.ru, megapolis.redbull@gmail.com, chshepkin@gkm-kz.com, golovin.sanya.71@gmail.com, shaihiev.i@outlook.com, TKmegapolis.zakaz@gkm-kz.com', 'ТОВАРИЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "KAZNORD"': 'buh3009@mail.ru'}

    print('emails:', emails)

    infotable = load_workbook(excel_file)
    info_sheet = infotable['Сводная таблица']

    for key, emails_ in emails.items():
        print('----------')
        print(key, suppliers_excels.get(key))
        emls = []

        row = None

        for i in range(2, info_sheet.max_row):
            print(info_sheet[f'A{i}'].value, key, info_sheet[f'A{i}'].value == key)
            if info_sheet[f'A{i}'].value == key:
                row = i
                break

        try:
            attachment = suppliers_excels.get(key)
            print(attachment)
            smtp_send(f'Рассылка для {key}', url=smtp_host, to=['Abdykarim.D@magnum.kz', 'Novitskaya@magnum.kz', 'Begaidarov@magnum.kz', 'Mukhtarova@magnum.kz'], subject=f'Исмет Рассылка Тест - {key}', username=smtp_author, attachments=[attachment])
            print(f"smtp_send('assdf', url=smtp_host, to={[email.strip() for email in emails_.split(',')]}, subject=f'Исмет Рассылка Тест', username=smtp_author)")
            info_sheet[f'C{row}'].value = 'Успешно отправлено'
        except Exception as error:
            print('ERROR', key)
            info_sheet[f'C{row}'].value = f'Ошибка при отправке - {error}'

    infotable.save(excel_file)
    sleep(10)
    smtp_send(f'Excel файл для Исмет Рассылки', url=smtp_host, to=['Abdykarim.D@magnum.kz', 'Novitskaya@magnum.kz', 'Begaidarov@magnum.kz', 'Mukhtarova@magnum.kz'], subject=f'Исмет Рассылка Тест', username=smtp_author, attachments=[excel_file])

    # smtp_send('assdf', url=smtp_host, to=['Abdykarim.D@magnum.kz'], subject=f'Исмет Рассылка Тест', username=smtp_author)



